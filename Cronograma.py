import glob
import os
import sys
import logging
import concurrent.futures
import base64
import random
import time
import subprocess
import multiprocessing
from datetime import datetime
import openpyxl
import requests
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
from tqdm import tqdm

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

COLUMN_MAP = {
    '24': 'Y',
    '25': 'Z',
    '26': 'AA', 
    '27': 'AB',
    '28': 'AC',
    '29': 'AD',
}

# Configuración de directorios
DOWNLOAD_DIR = 'downloads'
PROCESSED_DIR = 'processed'
# Calcular número máximo de hilos basado en CPUs disponibles
MAX_WORKERS = min((multiprocessing.cpu_count() * 2) + 4, 32)
logger.info(f"🖥️ Número de hilos configurados: {MAX_WORKERS}")

def setup_folders():
    """Crea las carpetas necesarias para el procesamiento"""
    # Obtener la ruta del escritorio
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    main_folder = os.path.join(desktop, "Resultados Python cronograma")
    excel_folder = os.path.join(main_folder, "hallazgos excel")
    html_folder = os.path.join(main_folder, "hallazgos html")
    
    # Crear la estructura de carpetas
    folders = [main_folder, excel_folder, html_folder, DOWNLOAD_DIR, PROCESSED_DIR, 'temp_images']
    for folder in folders:
        os.makedirs(folder, exist_ok=True)
    
    return excel_folder, html_folder

def process_images():
    """Redimensiona imágenes a 300x300 píxeles"""
    images = glob.glob(f'{DOWNLOAD_DIR}/*')
    logger.info(f"Procesando {len(images)} imágenes")
    
    with tqdm(total=len(images), desc="Procesando imágenes") as pbar:
        for path in images:
            try:
                img = PILImage.open(path)
                new_img = img.resize((300, 300))
                new_img.save(f'{PROCESSED_DIR}/{os.path.basename(path)}')
            except Exception as e:
                logger.error(f'Error procesando imagen {path}: {str(e)}')
            pbar.update(1)

def empty_folder(folder_name):
    """Vacía el contenido de una carpeta"""
    files = glob.glob(f'{folder_name}/*')
    for file in files:
        try:
            os.remove(file)
        except Exception as e:
            logger.error(f'Error eliminando {file}: {str(e)}')

def download_image(url, row, col):
    """Descarga una única imagen desde una URL con reintentos y timeouts progresivos."""
    max_retries = 3
    base_timeout = 10  # timeout base en segundos
    timeout_multiplier = 1.5  # multiplicador para cada reintento
    
    for attempt in range(max_retries):
        try:
            current_timeout = base_timeout * (timeout_multiplier ** attempt)
            response = requests.get(url, timeout=current_timeout)
            
            if not response.ok:
                logger.error(f"ERROR CON IMAGEN {row=} {col=} {url} - Código de estado: {response.status_code}")
                if attempt == max_retries - 1:
                    return False, None
                continue
                
            content_type = response.headers.get('content-type', '')
            if not content_type.startswith('image/'):
                logger.warning(f"NO ES UNA IMAGEN {row=} {col=} {url} - Tipo de contenido: {content_type}")
                return False, None
                
            extension = content_type.split('/')[-1].split(';')[0]
            if not extension:
                extension = 'jpg'  # Extensión por defecto si no se puede determinar
                
            path = f'{DOWNLOAD_DIR}/{row}_{col}.{extension}'
            
            with open(path, 'wb+') as f:
                f.write(response.content)
            
            # Convertir a base64 y almacenar en memoria
            img_base64 = base64.b64encode(response.content).decode('utf-8')
                
            logger.debug(f"Imagen descargada exitosamente: {path}")
            return True, img_base64
            
        except requests.exceptions.Timeout:
            logger.warning(f"Timeout al descargar imagen {row=} {col=} {url} (Intento {attempt + 1}/{max_retries})")
            if attempt == max_retries - 1:
                logger.error(f"Se agotaron los reintentos para la imagen {row=} {col=} {url}")
                return False, None
            continue
        except requests.exceptions.RequestException as e:
            logger.error(f"Error de red al descargar imagen {row=} {col=} {url}: {str(e)}")
            if attempt == max_retries - 1:
                return False, None
            continue
        except Exception as e:
            logger.error(f"Error inesperado al descargar imagen {row=} {col=} {url}: {str(e)}")
            return False, None

def download_images(sheet):
    """Descarga imágenes desde URLs en columnas 24-29 usando descargas paralelas."""
    download_tasks = []
    for row_index, row in enumerate(list(sheet.rows)[1:], 2):
        for col in range(24, 30):
            value = row[col].value
            if value and isinstance(value, str) and value.startswith('http'):
                download_tasks.append((value, row_index, col))
    
    if not download_tasks:
        logger.warning("No se encontraron URLs válidas para descargar")
        return {}
    
    completed = 0
    successful = 0
    images_base64 = {}  # Diccionario para almacenar imágenes en base64
    
    with tqdm(total=len(download_tasks), desc="Descargando imágenes") as pbar:
        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_task = {
                executor.submit(download_image, url, row, col): (url, row, col)
                for url, row, col in download_tasks
            }
            
            for future in concurrent.futures.as_completed(future_to_task):
                url, row, col = future_to_task[future]
                try:
                    success, img_base64 = future.result()
                    if success:  # Si la descarga fue exitosa
                        successful += 1
                        images_base64[f"{row}_{col}"] = img_base64
                except Exception as e:
                    logger.error(f"Error al descargar imagen en fila {row}, columna {col}: {e}")
                completed += 1
                pbar.set_postfix({'Exitosas': successful})
                pbar.update(1)
    
    logger.info(f"Descarga completada: {successful} de {len(download_tasks)} imágenes descargadas exitosamente")
    return images_base64

def add_images(sheet, filename):
    """Añade imágenes al archivo Excel"""
    for col in sheet.columns:
        sheet.column_dimensions[col[0].column_letter].width = 50
    for row in list(sheet.rows)[1:]:
        sheet.row_dimensions[row[0].row].height = 245

    images = glob.glob(f'{PROCESSED_DIR}/*')
    logger.info(f"Añadiendo {len(images)} imágenes al Excel")
    
    with tqdm(total=len(images), desc="Insertando imágenes") as pbar:
        for path in images:
            try:
                filename_parts = os.path.basename(path).split('.')[0].split('_')
                if len(filename_parts) == 2:
                    row, col = filename_parts
                    img = ExcelImage(path)
                    sheet.add_image(img, f'{COLUMN_MAP[col]}{row}')
            except Exception as e:
                logger.error(f'Error insertando imagen {path}: {str(e)}')
            pbar.update(1)
    sheet.parent.save(filename)

def create_sheet():
    """Crea una nueva hoja de cálculo"""
    wb = openpyxl.Workbook()
    wb.create_sheet('Export')
    del wb['Sheet']
    return wb['Export']

def filter_cobertura(sheet):
    """Filtra datos por columna de cobertura"""
    sheets = {}
    header = None
    for row in sheet.iter_rows(values_only=True):
        if not header:
            header = row
            continue
        if row[1]:
            if row[1] not in sheets:
                new_sheet = create_sheet()
                new_sheet.append(header)
                sheets[row[1]] = new_sheet
            sheets[row[1]].append(row)
    return sheets

def get_image_base64(url):
    """Convierte imagen a base64"""
    try:
        response = requests.get(url, timeout=10)
        if response.ok and response.headers.get('content-type', '').startswith('image/'):
            return base64.b64encode(response.content).decode('utf-8')
        return None
    except Exception as e:
        print(f'Error obteniendo imagen {url}: {str(e)}')
        return None

def extract_report_data(excel_file, images_base64):
    """Extrae datos para el reporte"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        report_data = {
            'filename': os.path.basename(excel_file),
            'sections': []
        }
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
            section_data = {
                'title': sheet_name,
                'entries': []
            }
            
            for row in sheet.iter_rows(min_row=2):
                entry = {}
                images = []
                
                for idx, cell in enumerate(row):
                    if idx >= len(headers):
                        continue
                    header = headers[idx]
                    value = cell.value
                    
                    # Excluir columnas específicas
                    if header in ['Conca-1', 'Total Horas', 'Recorridospedestres', 'Recuento de Combinada']:
                        continue
                    
                    if (cell.column in range(24, 30) and value and 
                        isinstance(value, str) and value.startswith('http')):
                        img_key = f"{cell.row}_{cell.column}"
                        if img_key in images_base64:
                            images.append({
                                'title': header,
                                'data': images_base64[img_key],
                                'position': f'Columna {cell.column_letter}'
                            })
                    else:
                        entry[header] = value
                
                if entry:
                    section_data['entries'].append({
                        'data': entry,
                        'images': images
                    })
            
            report_data['sections'].append(section_data)
        
        return report_data
    except Exception as e:
        print(f"Error procesando archivo {excel_file}: {str(e)}")
        return None

def generate_html_report(data, html_folder, cobertura):
    """Genera reporte HTML con visor modal de imágenes para una cobertura específica"""
    if not data:
        return
    
    html_content = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Informe de Hallazgos - {cobertura}</title>
    <link href="https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;500;700&display=swap" rel="stylesheet">
    <style>
        :root {{
            --color-background: #E0F0FF;
            --color-title: #003B5C;
            --color-subtitle: #0077B3;
            --color-card: #A4C8E1;
            --color-card-secondary: #dedede;
            --color-button: #005A8D;
            --shadow-sm: 0 2px 4px rgba(0,0,0,0.1);
            --shadow-md: 0 4px 6px rgba(0,0,0,0.1);
            --shadow-lg: 0 10px 15px rgba(0,0,0,0.1);
            --border-radius: 8px;
            --spacing-xs: 0.25rem;
            --spacing-sm: 0.5rem;
            --spacing-md: 0.75rem;
            --spacing-lg: 2.5rem;
            --spacing-xl: 1.5rem;
        }}
        
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Roboto', sans-serif;
            line-height: 1.4;
            color: var(--color-title);
            background-color: var(--color-background);
            padding: var(--spacing-sm);
        }}
        
        .report-container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: var(--border-radius);
            box-shadow: var(--shadow-lg);
            overflow: hidden;
        }}
        
        header {{
            background: var(--color-title);
            color: white;
            padding: var(--spacing-md);
            text-align: center;
            position: relative;
            overflow: hidden;
        }}
        
        header::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: linear-gradient(45deg, var(--color-title), var(--color-subtitle));
            opacity: 0.9;
            z-index: 0;
        }}
        
        header h1 {{
            font-size: 1.8rem;
            margin-bottom: var(--spacing-xs);
            position: relative;
            z-index: 1;
        }}
        
        header p {{
            color: rgba(255, 255, 255, 0.9);
            font-size: 0.9rem;
            position: relative;
            z-index: 1;
        }}
        
        .report-section {{
            padding: var(--spacing-md);
            border-bottom: 1px solid rgba(0, 0, 0, 0.1);
        }}
        
        .report-section:last-child {{
            border-bottom: none;
        }}
        
        .report-section h3 {{
            color: var(--color-subtitle);
            font-size: 1.2rem;
            margin-bottom: var(--spacing-md);
            padding-bottom: var(--spacing-xs);
            border-bottom: 2px solid var(--color-subtitle);
        }}
        
        .entry {{
            background: var(--color-card-secondary);
            border-radius: var(--border-radius);
            margin-bottom: var(--spacing-xl);
            overflow: hidden;
            box-shadow: var(--shadow-sm);
            padding: var(--spacing-md);
        }}
        
        .entry:last-child {{
            margin-bottom: 0;
        }}
        
        .entry-divider {{
            height: 2px;
            background: linear-gradient(to right, transparent, var(--color-subtitle), transparent);
            margin: var(--spacing-lg) 0;
            opacity: 0.5;
        }}
        
        .data-flex {{
            display: flex;
            flex-wrap: wrap;
            gap: var(--spacing-md);
            margin-bottom: var(--spacing-md);
        }}
        
        .data-item {{
            flex: 1 1 300px;
            min-width: 300px;
            background: white;
            padding: var(--spacing-sm);
            border-radius: var(--border-radius);
            transition: transform 0.2s ease;
        }}
        
        .data-item:hover {{
            transform: translateY(-2px);
        }}
        
        .data-item-label {{
            font-size: 0.75rem;
            color: var(--color-subtitle);
            margin-bottom: var(--spacing-xs);
            font-weight: 500;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        
        .data-item-value {{
            font-size: 0.85rem;
            color: var(--color-title);
            word-break: break-word;
        }}
        
        .data-item-full {{
            flex: 1 1 100%;
            background: linear-gradient(to right, white, var(--color-background));
        }}
        
        .image-gallery {{
            display: flex;
            flex-wrap: wrap;
            gap: var(--spacing-md);
            margin-top: var(--spacing-md);
            padding: var(--spacing-sm);
        }}
        
        .image-item {{
            flex: 1 1 250px;
            min-width: 250px;
            background: white;
            border-radius: var(--border-radius);
            overflow: hidden;
            box-shadow: var(--shadow-sm);
            transition: transform 0.2s ease;
        }}
        
        .image-item:hover {{
            transform: translateY(-3px);
            box-shadow: var(--shadow-md);
        }}
        
        .image-container {{
            height: 200px;
            display: flex;
            align-items: center;
            justify-content: center;
            background: var(--color-background);
            cursor: pointer;
            position: relative;
            overflow: hidden;
        }}
        
        .image-container::after {{
            content: '🔍';
            position: absolute;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            font-size: 1.2rem;
            opacity: 0;
            transition: opacity 0.3s ease;
            background: rgba(0, 0, 0, 0.5);
            width: 100%;
            height: 100%;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
        }}
        
        .image-container:hover::after {{
            opacity: 1;
        }}
        
        .image-container img {{
            max-width: 100%;
            max-height: 100%;
            object-fit: contain;
        }}
        
        .image-caption {{
            padding: var(--spacing-sm);
            background: white;
            font-size: 0.8rem;
            color: var(--color-subtitle);
            text-align: center;
        }}
        
        /* Modal styles */
        .modal {{
            display: none;
            position: fixed;
            z-index: 1000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background: rgba(0, 0, 0, 0.9);
            backdrop-filter: blur(5px);
        }}
        
        .modal-content {{
            display: block;
            max-width: 90%;
            max-height: 90%;
            position: absolute;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            border-radius: var(--border-radius);
            box-shadow: var(--shadow-lg);
        }}
        
        .close {{
            position: absolute;
            top: 20px;
            right: 30px;
            color: white;
            font-size: 40px;
            font-weight: bold;
            cursor: pointer;
            transition: color 0.3s ease;
        }}
        
        .close:hover {{
            color: var(--color-button);
        }}
        
        /* Responsive Design */
        @media (max-width: 1200px) {{
            .data-item {{
                flex: 1 1 250px;
                min-width: 250px;
            }}
        }}
        
        @media (max-width: 992px) {{
            .data-item {{
                flex: 1 1 200px;
                min-width: 200px;
            }}
            
            .image-item {{
                flex: 1 1 200px;
                min-width: 200px;
            }}
            
            .image-container {{
                height: 180px;
            }}
        }}
        
        @media (max-width: 768px) {{
            body {{
                padding: var(--spacing-xs);
            }}
            
            .report-container {{
                border-radius: 0;
            }}
            
            header h1 {{
                font-size: 1.5rem;
            }}
            
            .report-section h3 {{
                font-size: 1.1rem;
            }}
            
            .entry {{
                margin-bottom: var(--spacing-lg);
                padding: var(--spacing-sm);
            }}
            
            .data-item,
            .image-item {{
                flex: 1 1 30%;
                min-width: 48%;
            }}
            
            .image-container {{
                height: 150px;
            }}
        }}
        
        @media (max-width: 480px) {{
            header h1 {{
                font-size: 1.3rem;
            }}
            
            header p {{
                font-size: 0.8rem;
            }}
            
            .report-section h3 {{
                font-size: 1rem;
            }}
            
            .data-item-label {{
                font-size: 0.7rem;
            }}
            
            .data-item-value {{
                font-size: 0.8rem;
            }}
            
            .image-container {{
                height: 200px;
            }}
        }}
    </style>
</head>
<body>
    <div class="report-container">
        <header>
            <h1>Informe de Hallazgos</h1>
            <p>{cobertura} - Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}</p>
        </header>
"""

    for section in data['sections']:
        html_content += f"""
        <div class="report-section">
            <h3>{section['title']}</h3>"""
        
        for entry in section['entries']:
            html_content += """
            <div class="entry">
                <div class="data-flex">"""
            
            # Lista de campos que deberían ocupar todo el ancho
            full_width_fields = ['ObservacionesHallazgo']
            
            for key, value in entry['data'].items():
                if value and str(value).strip():
                    css_class = 'data-item-full' if key in full_width_fields else ''
                    html_content += f"""
                    <div class="data-item {css_class}">
                        <div class="data-item-label">{key}</div>
                        <div class="data-item-value">{value}</div>
                    </div>"""
            
            html_content += """
                </div>"""
            
            if entry['images']:
                html_content += """
                <div class="image-gallery">"""
                for img in entry['images']:
                    html_content += f"""
                    <div class="image-item">
                        <div class="image-container" onclick="openModal('data:image/png;base64,{img['data']}')">
                            <img src="data:image/png;base64,{img['data']}" alt="{img['title']}">
                        </div>
                        <div class="image-caption">
                            {img['title']} ({img['position']})
                        </div>
                    </div>"""
                html_content += """
                </div>"""
        
            html_content += """
            </div>"""
            html_content += """
            <div class="entry-divider"></div>"""
        html_content += """
        </div>"""
    
    html_content += """
        <!-- Modal para imágenes -->
        <div id="imageModal" class="modal">
            <span class="close" onclick="closeModal()">&times;</span>
            <img class="modal-content" id="modalImage">
        </div>
        
        <script>
            function openModal(imgSrc) {
                document.getElementById('modalImage').src = imgSrc;
                document.getElementById('imageModal').style.display = 'block';
                document.body.style.overflow = 'hidden';
            }
            
            function closeModal() {
                document.getElementById('imageModal').style.display = 'none';
                document.body.style.overflow = 'auto';
            }
            
            window.onclick = function(event) {
                if (event.target == document.getElementById('imageModal')) {
                    closeModal();
                }
            }
            
            document.addEventListener('keydown', function(event) {
                if (event.key === 'Escape') {
                    closeModal();
                }
            });
        </script>
    </div>
</body>
</html>"""

    report_filename = os.path.join(html_folder, f"informe_{cobertura}.html")
    with open(report_filename, "w", encoding="utf-8") as f:
        f.write(html_content)
    
    print(f"\n✅ Informe generado: {report_filename}")

def clean_temp_files():
    """Elimina los archivos y carpetas temporales, dejando solo la carpeta de output"""
    temp_folders = [DOWNLOAD_DIR, PROCESSED_DIR, 'temp_images']
    total_files = 0
    
    # Primero contar todos los archivos
    for folder in temp_folders:
        if os.path.exists(folder):
            files = glob.glob(f'{folder}/*')
            total_files += len(files)
    
    if total_files == 0:
        logger.info('✅ No hay archivos temporales para limpiar')
        return
    
    logger.info(f'🧹 Limpiando {total_files} archivos temporales...')
    processed_files = 0
    
    with tqdm(total=total_files, desc="Limpiando archivos temporales") as pbar:
        for folder in temp_folders:
            try:
                if os.path.exists(folder):
                    files = glob.glob(f'{folder}/*')
                    for file in files:
                        try:
                            os.remove(file)
                            processed_files += 1
                        except Exception as e:
                            logger.error(f'Error eliminando archivo {file}: {str(e)}')
                        pbar.update(1)
                    try:
                        os.rmdir(folder)
                        logger.info(f'✅ Carpeta temporal eliminada: {folder}')
                    except Exception as e:
                        logger.error(f'Error eliminando carpeta {folder}: {str(e)}')
            except Exception as e:
                logger.error(f'Error limpiando carpeta {folder}: {str(e)}')
    
    logger.info(f'✅ Limpieza completada: {processed_files} de {total_files} archivos eliminados')

def main():
    """Función principal"""
    excel_folder, html_folder = setup_folders()
    
    try:
        # Solicitar la ruta del archivo Excel
        while True:
            file_path = input("Por favor, ingrese la ruta completa del archivo Excel: ").strip()
            if not file_path:
                logger.error("❌ No se proporcionó ninguna ruta")
                continue
            if not os.path.exists(file_path):
                logger.error(f"❌ El archivo no existe: {file_path}")
                continue
            if not file_path.lower().endswith('.xlsx'):
                logger.error("❌ El archivo debe tener extensión .xlsx")
                continue
            break

        wb = openpyxl.load_workbook(file_path)
        logger.info(f"📂 Archivo cargado: {file_path}")
        
        sheet = wb['Export']
        sheets = filter_cobertura(sheet)
        output_files = []
        
        with tqdm(total=len(sheets), desc="Procesando coberturas") as pbar:
            for cobertura, sheet in sheets.items():
                logger.info(f"\n🔄 Procesando: {cobertura}")
                empty_folder(DOWNLOAD_DIR)
                empty_folder(PROCESSED_DIR)
                
                # Descargar imágenes y obtener base64
                images_base64 = download_images(sheet)
                process_images()
                
                output_filename = os.path.join(excel_folder, f'{cobertura}.xlsx')
                logger.info(f"💾 Guardando: {output_filename}")
                add_images(sheet, output_filename)
                output_files.append((output_filename, cobertura, images_base64))
                
                # Espera aleatoria entre coberturas
                if len(sheets) > 1:  # Solo esperar si hay más de una cobertura
                    base_wait = 5  # tiempo base en segundos
                    max_wait = 12  # tiempo máximo en segundos
                    wait_time = min(base_wait * (1.5 ** (len(output_files) - 1)), max_wait)
                    wait_time = random.uniform(wait_time * 0.8, wait_time)  # Variación aleatoria
                    logger.info(f"⏳ Esperando {wait_time:.1f} segundos antes de procesar la siguiente cobertura...")
                    time.sleep(wait_time)
                
                pbar.update(1)
        
        logger.info("\n✅ Proceso completado exitosamente")
        
        # Generar reportes HTML
        if output_files:
            logger.info("📊 Generando reportes HTML...")
            with tqdm(total=len(output_files), desc="Generando reportes HTML") as pbar:
                for excel_file, cobertura, images_base64 in output_files:
                    data = extract_report_data(excel_file, images_base64)
                    if data:
                        generate_html_report(data, html_folder, cobertura)
                    pbar.update(1)
        
        # Limpiar archivos temporales
        logger.info("\n🧹 Limpiando archivos temporales...")
        clean_temp_files()
        
        # Abrir la carpeta de resultados
        main_folder = os.path.dirname(excel_folder)
        subprocess.Popen(f'explorer "{main_folder}"')
        logger.info(f"📂 Abriendo carpeta de resultados: {main_folder}")
        
    except Exception as e:
        logger.error(f"❌ Error: {str(e)}")
        sys.exit(1)

if __name__ == '__main__':
    main()